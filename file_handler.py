import PyPDF2
import openpyxl

def extraer_texto_pdf(ruta: str) -> str:
    texto = ""
    with open(ruta, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for pagina in reader.pages:
            texto += pagina.extract_text() or ""
    return texto

def extraer_texto_excel(ruta: str) -> str:
    wb = openpyxl.load_workbook(ruta)
    texto = ""
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        texto += f"\nHoja: {hoja}\n"
        for fila in ws.iter_rows(values_only=True):
            fila_limpia = [str(c) for c in fila if c is not None]
            if fila_limpia:
                texto += ", ".join(fila_limpia) + "\n"
    return texto